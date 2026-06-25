from __future__ import annotations

import argparse
import os
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


LEVEL_NAMES = {
    0: "Plant / Unit",
    1: "System",
    2: "Equipment unit",
    3: "Equipment component",
}


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def make_record(
    *,
    kks_code: str,
    description: str,
    level_no: int,
    plant_code: str,
    system_code: str = "",
    equipment_code: str = "",
    component_code: str = "",
    responsible_area: str = "",
    source_kind: str,
    source_row: int | None = None,
) -> dict[str, object]:
    return {
        "kks_code": kks_code,
        "description": description,
        "hierarchy_level": LEVEL_NAMES[level_no],
        "level_no": level_no,
        "plant_code": plant_code,
        "system_code": system_code,
        "equipment_code": equipment_code,
        "component_code": component_code,
        "responsible_area": responsible_area,
        "source_kind": source_kind,
        "source_row": source_row,
    }


def parent_code(record: dict[str, object]) -> str | None:
    level = int(record["level_no"])
    if level == 0:
        return None
    if level == 1:
        return str(record["plant_code"])
    if level == 2:
        return f"{record['plant_code']} {record['system_code']}".strip()
    return (
        f"{record['plant_code']} {record['system_code']}{record['equipment_code']}"
    ).strip()


def read_workbook(path: Path) -> tuple[dict[str, dict[str, object]], list[dict], int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "kks" not in workbook.sheetnames:
        raise ValueError("Workbook must contain a sheet named 'kks'")
    sheet = workbook["kks"]
    rows = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    required = {
        "PLANT CODE",
        "SYSTEM CODE",
        "EQUIPMENT CODE",
        "COMPONENT CODE",
        "DESCRIPTION",
        "KKS code",
        "RESPONSIBLE AREA",
    }
    missing = required.difference(headers)
    if missing:
        raise ValueError(f"Missing workbook columns: {', '.join(sorted(missing))}")

    positions = {name: headers.index(name) for name in required}
    records: dict[str, dict[str, object]] = {}
    issues: list[dict] = []
    source_rows = 0

    for source_row, values in enumerate(rows, start=2):
        source_rows += 1
        plant = clean(values[positions["PLANT CODE"]])
        system = clean(values[positions["SYSTEM CODE"]])
        equipment = clean(values[positions["EQUIPMENT CODE"]])
        component = clean(values[positions["COMPONENT CODE"]])
        description = clean(values[positions["DESCRIPTION"]])
        kks_code = clean(values[positions["KKS code"]])
        area = clean(values[positions["RESPONSIBLE AREA"]])

        if not kks_code:
            issues.append(
                {
                    "source_row": source_row,
                    "issue_type": "blank_kks",
                    "kks_code": "",
                    "description": description,
                    "responsible_area": area,
                    "details": "Row skipped because KKS code is blank.",
                }
            )
            continue

        level_no = 3 if component else 2 if equipment else 1
        record = make_record(
            kks_code=kks_code,
            description=description or "Description not supplied",
            level_no=level_no,
            plant_code=plant,
            system_code=system,
            equipment_code=equipment,
            component_code=component,
            responsible_area=area,
            source_kind="kks_workbook",
            source_row=source_row,
        )
        if kks_code in records:
            issues.append(
                {
                    "source_row": source_row,
                    "issue_type": "duplicate_kks",
                    "kks_code": kks_code,
                    "description": description,
                    "responsible_area": area,
                    "details": (
                        "First occurrence retained from source row "
                        f"{records[kks_code]['source_row']}."
                    ),
                }
            )
            continue
        records[kks_code] = record

    workbook.close()
    return records, issues, source_rows


def add_hierarchy_parents(records: dict[str, dict[str, object]]) -> int:
    inferred_count = 0
    source_records = list(records.values())
    plants = sorted({str(record["plant_code"]) for record in source_records})
    for plant in plants:
        if plant not in records:
            records[plant] = make_record(
                kks_code=plant,
                description=f"Morupule B plant area {plant}",
                level_no=0,
                plant_code=plant,
                source_kind="kks_inferred",
            )
            inferred_count += 1

    for record in source_records:
        level = int(record["level_no"])
        if level >= 2:
            system_kks = f"{record['plant_code']} {record['system_code']}".strip()
            if system_kks not in records:
                records[system_kks] = make_record(
                    kks_code=system_kks,
                    description="System parent inferred from KKS register",
                    level_no=1,
                    plant_code=str(record["plant_code"]),
                    system_code=str(record["system_code"]),
                    responsible_area=str(record["responsible_area"]),
                    source_kind="kks_inferred",
                )
                inferred_count += 1
        if level == 3:
            equipment_kks = (
                f"{record['plant_code']} {record['system_code']}"
                f"{record['equipment_code']}"
            ).strip()
            if equipment_kks not in records:
                records[equipment_kks] = make_record(
                    kks_code=equipment_kks,
                    description="Equipment parent inferred from KKS register",
                    level_no=2,
                    plant_code=str(record["plant_code"]),
                    system_code=str(record["system_code"]),
                    equipment_code=str(record["equipment_code"]),
                    responsible_area=str(record["responsible_area"]),
                    source_kind="kks_inferred",
                )
                inferred_count += 1
    return inferred_count


def import_records(
    database,
    source_file: Path,
    records: dict[str, dict[str, object]],
    issues: list[dict],
    source_rows: int,
    inferred_count: int,
) -> None:
    source_name = str(source_file.resolve())
    imported_at = datetime.now().isoformat(timespec="seconds")
    database.execute("DELETE FROM kks_import_issues WHERE source_file = ?", (source_name,))

    existing = {
        row["kks_code"]: (row["id"], bool(row["is_reference"]))
        for row in database.execute(
            "SELECT id, kks_code, is_reference FROM assets"
        ).fetchall()
    }

    for level in range(4):
        inserts = []
        updates = []
        for record in records.values():
            if int(record["level_no"]) != level:
                continue
            code = str(record["kks_code"])
            parent = parent_code(record)
            parent_id = existing.get(parent, (None, False))[0] if parent else None
            values = (
                parent_id,
                code,
                record["description"],
                record["hierarchy_level"],
                record["level_no"],
                record["plant_code"],
                record["system_code"],
                record["equipment_code"],
                record["component_code"],
                record["responsible_area"],
                record["source_kind"],
                record["source_row"],
            )
            if code not in existing:
                inserts.append(values)
            elif not existing[code][1]:
                updates.append(values)

        database.executemany(
            """
            INSERT INTO assets (
                parent_id, kks_code, description, hierarchy_level, level_no,
                plant_code, system_code, equipment_code, component_code,
                responsible_area, source_kind, source_row
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            inserts,
        )
        database.executemany(
            """
            UPDATE assets SET
                parent_id = ?, description = ?, hierarchy_level = ?, level_no = ?,
                plant_code = ?, system_code = ?, equipment_code = ?,
                component_code = ?, responsible_area = ?, source_kind = ?,
                source_row = ?, status = 'active'
            WHERE kks_code = ?
            """,
            [
                (
                    values[0], values[2], values[3], values[4], values[5],
                    values[6], values[7], values[8], values[9], values[10],
                    values[11], values[1],
                )
                for values in updates
            ],
        )
        existing = {
            row["kks_code"]: (row["id"], bool(row["is_reference"]))
            for row in database.execute(
                "SELECT id, kks_code, is_reference FROM assets"
            ).fetchall()
        }

    database.executemany(
        """
        INSERT INTO kks_import_issues (
            source_file, source_row, issue_type, kks_code, description,
            responsible_area, details, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            (
                source_name,
                issue["source_row"],
                issue["issue_type"],
                issue["kks_code"],
                issue["description"],
                issue["responsible_area"],
                issue["details"],
                imported_at,
            )
            for issue in issues
        ],
    )
    source_unique = sum(
        record["source_kind"] == "kks_workbook" for record in records.values()
    )
    duplicate_rows = sum(
        issue["issue_type"] == "duplicate_kks" for issue in issues
    )
    database.execute(
        """
        INSERT INTO kks_import_runs (
            source_file, imported_at, source_rows, unique_assets,
            duplicate_rows, inferred_parents
        ) VALUES (?, ?, ?, ?, ?, ?)
        """,
        (
            source_name,
            imported_at,
            source_rows,
            source_unique,
            duplicate_rows,
            inferred_count,
        ),
    )
    database.commit()


def main() -> None:
    parser = argparse.ArgumentParser(description="Import Morupule B KKS workbook.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument(
        "--data-dir",
        type=Path,
        help="Runtime data directory containing morupule_sipam.db and uploads.",
    )
    parser.add_argument(
        "--database",
        type=Path,
        help="Explicit SQLite database path. Overrides --data-dir for the database only.",
    )
    arguments = parser.parse_args()
    if arguments.data_dir:
        os.environ["SIPAM_DATA_DIR"] = str(arguments.data_dir.resolve())
    if arguments.database:
        os.environ["SIPAM_DATABASE"] = str(arguments.database.resolve())

    from app import app, get_db, init_db

    workbook_path = arguments.workbook.resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)

    records, issues, source_rows = read_workbook(workbook_path)
    inferred_count = add_hierarchy_parents(records)
    with app.app_context():
        init_db()
        import_records(
            get_db(),
            workbook_path,
            records,
            issues,
            source_rows,
            inferred_count,
        )

    issue_counts = Counter(issue["issue_type"] for issue in issues)
    source_unique = sum(
        record["source_kind"] == "kks_workbook" for record in records.values()
    )
    print(f"Source rows: {source_rows:,}")
    print(f"Unique workbook assets: {source_unique:,}")
    print(f"Inferred hierarchy nodes: {inferred_count:,}")
    print(f"Duplicate rows audited: {issue_counts['duplicate_kks']:,}")


if __name__ == "__main__":
    main()
